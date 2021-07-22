import os
import sys
import csv
import json
import time
import numpy
import base64
import nibabel
import logging
import openpyxl
import platform
from pydicom.dataset import Dataset
from pydicom.filereader import dcmread
from matplotlib import pyplot as plt
from matplotlib import use as matuse
from nibabel.affines import apply_affine
from openpyxl.styles import Alignment
from openpyxl.utils.exceptions import InvalidFileException
from PyQt6.QtCore import QObject, QRegularExpression, QThread, pyqtBoundSignal, pyqtSignal, Qt, pyqtSlot
from PyQt6.QtGui import QCursor, QDoubleValidator, QIcon, QMouseEvent, QPixmap, QRegularExpressionValidator
from PyQt6.QtWidgets import QApplication, QCheckBox, QComboBox, QDialog, QFileDialog, QGridLayout, QHBoxLayout, QLabel, QLineEdit, QMainWindow, QMessageBox, QPlainTextEdit, QProgressBar, QPushButton, QSystemTrayIcon, QVBoxLayout,QWidget

os.chdir(os.path.split(os.path.realpath(__file__))[0])
# 更改工作目录到脚本文件夹，使相对路径均以脚本文件夹为基准
logging.basicConfig(format="%(asctime)s-%(levelname)s-%(message)s",datefmt="%Y-%m-%d %H:%M:%S")
# logging的基本设置
APPID="fMRI Process Tools"
ICON='''iVBORw0KGgoAAAANSUhEUgAAAGAAAABgCAYAAADimHc4AAAD6UlEQVR4AWIYBaMA0N49wMqVhgEYzvWtbUQbrG3btq2oWNu7wdq2vVubF0Ftm0FtfH3rXp5zZn6153uTJ04uZ+bnjKZ
        pmqZpmqZp2r4dhjuzdDgySDsdmyFZ2oLzkDDtE4gh3yFhWn+IIaVImDYbYshSJEgrwhaIQY0RM+0QiGHHIWbaFRDDbkLMtK4Qw55FzLSPIYZ9i5hpfSGGDUfMtJkQwxbjgK4Yn6A/Hk
        NLJC0XN2MzxIK7kI+kNUNX9MOnqI/geg+yjw34FecgqjzcikkQy6bj7ph/iNPxE9ZB9vFFiOP2jZAaTMEjaI7KHYXREMcm4gRUrik6Y2LEAt+xCKbekBjW42eciXy8gI0QTzbhDRTiV
        HyPdZAYhiGILoNkYCUkECv318ldAaZAUmou6nqfsabcS/BSCyyHpNxadITzPoOoHX6D046sMFlSW3E6nDUIoioYiVxY71qIqtb9sFoRZgbwcJ+OQfhxl0GYjq0Qj5agEaz1FMSTvrgP
        rVBTrXAf+kI8eQtWaopVEMdKcCqSdipKII5tsDUsPR/i0GY8gmx7xMOI7UoYrwGmOVybuRimutjh2tN8NIGVmqOfg//8i2G6ix08EsrRHlbLw2sWRxyPwFaPQCz5CIVw1uUW1oJKYLs
        SiEFrcDu8dBDGQAw5FbY7FWLINBwBr9XF95As9YGr+kCy9B8aIZgexgZIhu6Dq+7LcpDwNHIQXK9kcYulFVzVKovT1x8j2L6EZGAqXDcVkoHfEWw9IRkYBNcNCvZ79TDE+xmu+xmSgT
        IEW5n+Afw2UJ+C/PZbCl6E/0awfZyCYeg3CK4cPI3NKZiIbcELyEUQNcJ/KVyK6BHC1dcjMC3Fi3HTfS7G3Y410OVoLpfAWYX4CKIbMhW8j3xYrT3KdUuyRsPQBlZqgvm6KR9ptq1N+
        Sv1WEpsl8F4HbFBD2ZFWovWsNJbejQx0iu2J11L9HBujeahHqx2P0RV61ZYLxcjIaqCEuTASadXeLirLTjO39q/+hrO64i1kJRbidbw0kuQlHsc3qqLuZCUmoZCeO2mFL9ZxxUIomGQ
        GNbhe5yKQryBTRBPNuJFFOBM/Iz18Wfm4XRsxCb3RHRGU1TuBEyEODYOx9RwC+iRiHeB2YRDEVRfVPPf/lPMK/z5uBvTIZZNxh3IR1Tn4NdqFiHfQ3DVx6foh65ohqTl4y6Ly9u3IA9
        Ja4nH0B+foBgHdIshhs2EFrPhlpaztZh9C4lm62KF9izEsK7QLEzuLEyatOMghh0CLWaNLazXF0FL0FLDZ3W0hJVCDOkPj+lQ9BNoCTvb4EcZngXv6Yd5apqmaZqmaZqmado2q+hsvC
        wTK/oAAAAASUVORK5CYII=''' 
# 预定义的常量，依次为程序名，程序图标的base64编码结果
if platform.system()=="Windows":
    import ctypes
    ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID(APPID)
    # 让Windows的任务栏图标可以正常显示
matuse("Agg")
# 让matplotlib使用Agg后端避免Tkinter在非主线程运行的问题

class Processor(QObject):
    def __init__(self,infopath:str,update_progress_signal:pyqtBoundSignal,busy_progress:pyqtBoundSignal,finish_signal:pyqtBoundSignal):
        super().__init__()
        self.logger=logging.getLogger()
        self.update_progress_signal=update_progress_signal
        self.busy_progress=busy_progress
        self.finish_signal=finish_signal
        with open(infopath,"r",encoding="utf-8") as reader:
            self.config=json.load(reader)
        for key in ["split","newline"]:
            self.config[key]=self.config[key].replace("\\n","\n").replace("\\t","\t")
        self.logger.debug("已加载处理信息文件，配置信息：%s" %self.config)
    def csv_to_excel(self,csvpath:str):
        sheet=openpyxl.Workbook().active
        if self.config["node_start"]>=1:
            sheet.insert_rows(1,self.config["node_start"])
        with open(csvpath,"r",newline="") as csvfile:
            [sheet.append(row) for row in csv.reader(csvfile)]
        return sheet
    def is_text(self,path:str):
        size=512
        if os.path.isdir(path):
            return False
        try:
            with open(path,"tr") as reader:
                reader.read(size)
        except:
            return False
        else:
            return True
    def get_common_part(self,datas:list,overnum:int):
        max_size=max([matrix.shape for matrix in datas])
        if all([matrix.shape==max_size for matrix in datas])==False:
            self.logger.warning("存在一些大小偏小的矩阵，我们将以0补齐使得所有矩阵大小一致")
        common=numpy.zeros(max_size,float)
        for index,value in numpy.ndenumerate(common):
            count=0
            for data in datas:
                if data[index]!=0:
                    count+=1
            self.logger.debug("坐标(%d,%d)存在 %d 次" %(index[0],index[1],count))
            if count>=overnum:
                if value!=0:
                    self.logger.warning("公共部分矩阵在(%d,%d)似乎出现了非0初始化值 %d？" %(index[0],index[1],value))
                common[index]=count
        return common
    def handle_ds(self,ds:Dataset):
        '''
        功能:从DICOM文件中获取患者信息
        参数:ds:Dataset:pydicom打开文件后获得的数据集
        返回值:文件名、患者ID、患者姓名、患者生日、患者年龄、患者性别、患者身高、患者体重组成的列表以及除文件名以外的元素组成的列表；两个列表按顺序以元组形式返回
        注意:具体可以获取的内容取决于文件自身储存的信息，可以获取的项目名称需要使用相应程序（Photoshop 2021等）打开DICOM文件后在元数据处查看
        '''
        sexMap={"F":"女","M":"男"}
        funpath=str(ds.filename)
        patientId=str(ds.get("PatientID"))
        patientName=str(ds.get("PatientName"))
        patientBirthDate=str(ds.get("PatientBirthDate"))
        patientAge=str(ds.get("PatientAge"))
        patientSex=str(ds.get("PatientSex"))
        patientSize=str(ds.get("PatientSize"))
        patientWeight=str(ds.get("PatientWeight"))
        data=[funpath,patientId,patientName,patientBirthDate,patientAge,patientSex,patientSize,patientWeight]
        datab=data.copy()
        datab.remove(funpath)
        datab[1]=datab[1].title()
        # 格式化人名为单词首字母大写的形式
        datab[2]=time.strftime("%Y-%m-%d",time.strptime(datab[2],"%Y%m%d"))
        # 格式化时间为年-月-日的形式
        datab[3]="%d" %int(datab[3].replace("Y",""))
        # 格式化年龄为正常整数
        datab[4]=sexMap[datab[4]]
        # 格式化性别为正常名称
        if datab[5]=="None":
            datab[5]="无数据"
        # 标注无身高数据的项目
        return data,datab
    def main(self):
        start_time=time.time()
        self.update_progress_signal.emit(0)
        datas=list()
        workdir=self.config["workdir"]
        if workdir=="" or os.path.isfile(workdir):
            workdir=os.path.split(os.path.realpath(__file__))[0]
        if os.path.exists(os.path.join(workdir,"Results"))==False:
            os.mkdir(os.path.join(workdir,"Results"))
        files=[]
        [files.append(file) for file in os.listdir(workdir) if self.is_text(os.path.join(workdir,file))]
        # 跳过文件夹和非纯文本文件
        self.logger.debug("文件列表：%s" %files)
        try:
            nodepath=self.config["nodepath"]
            if nodepath.endswith(".xlsx"):
                nodes=openpyxl.load_workbook(nodepath).active
            else:
                nodes=self.csv_to_excel(nodepath)
        except InvalidFileException:
            self.logger.error("输入的节点信息不是正确的xlsx表格")
            self.finish_signal.emit(1)
        except:
            self.logger.error("输入节点信息不是正确的csv格式数据")
            self.finish_signal.emit(1)
        else:
            nodesmap={}
            self.logger.debug("加载节点信息 %s 成功" %os.path.basename(nodepath))
            results=openpyxl.Workbook()
            center_style=Alignment("center","center",wrap_text=True)
            try:
                target=self.config["target"]
                overnum=self.config["overnum"]
                selected_nodes=self.config["selected_nodepath"]
            except KeyError:
                self.logger.error("处理信息文件有误")
                self.finish_signal.emit(1)
            else:
                self.logger.debug("皮尔逊系数目标值：%f，最少图片数：%d" %(target,overnum))
                if target<=1 and target>=0 and overnum>0 and overnum<=len(files):
                    self.logger.debug("数据符合要求")
                    for file in files:
                        self.logger.info("正在处理文件 %s" %file)
                        try:
                            matrix=numpy.loadtxt(os.path.join(workdir,file))
                        except Exception as e:
                            self.logger.debug("解析文件 %s 出错，原因：%s" %(file,e))
                        else:
                            if matrix.size==0:
                                self.logger.debug("矩阵大小为空，将跳过")
                            else:
                                self.logger.debug("解析文件 %s 成功，矩阵大小：(%d,%d)" %(file,matrix.shape[0],matrix.shape[1]))
                                lines1=list()
                                lines2=list()
                                if file in results.sheetnames:
                                    sheet=results[file]
                                else:
                                    sheet=results.active
                                    sheet.title=file
                                sheetmap={
                                    "A1":"x",
                                    "B1":"y",
                                    "C1":"皮尔逊系数",
                                    "D1":"X节点结果",
                                    "D2":"x",
                                    "E2":"y",
                                    "F2":"z",
                                    "G2":"区域英文名",
                                    "H2":"区域中文名",
                                    "I1":"Y节点结果",
                                    "I2":"x",
                                    "J2":"y",
                                    "K2":"z",
                                    "L2":"区域英文名",
                                    "M2":"区域中文名"
                                    }
                                for sheetstr in sheetmap.keys():
                                    sheet[sheetstr]=sheetmap[sheetstr]
                                    sheet[sheetstr].alignment=center_style
                                for sheetstr in ["A1:A2","B1:B2","C1:C2","D1:H1","I1:M1"]:
                                    sheet.merge_cells(sheetstr)
                                # 绘制表头
                                for index,value in numpy.ndenumerate(matrix):
                                    if abs(value)>=target:
                                        r=self.config["node_start"]
                                        rdata={"A":index[0]+1,"B":index[1]+1,"C":value}
                                        for row in nodes.iter_rows(min_row=self.config["node_start"]):
                                            node_name="-"
                                            node_num=list(row)[0].value
                                            if nodes.cell(row=r,column=5).value!=None and self.config["add_label"]=="E":
                                                node_name=str(nodes.cell(row=r,column=5).value)
                                            elif nodes.cell(row=r,column=6).value!=None and self.config["add_label"]=="C":
                                                node_name=str(nodes.cell(row=r,column=6).value)
                                            if node_num!=None and int(node_num)==int(rdata["A"]):
                                                rdata["D"]=nodes.cell(row=r,column=2).value
                                                rdata["E"]=nodes.cell(row=r,column=3).value
                                                rdata["F"]=nodes.cell(row=r,column=4).value
                                                rdata["G"]=nodes.cell(row=r,column=5).value
                                                rdata["H"]=nodes.cell(row=r,column=6).value
                                                line1=self.config["split"].join([str(rdata["D"]),str(rdata["E"]),str(rdata["F"])])+self.config["split"]+"1"+self.config["split"]+"1.000000"+self.config["split"]+node_name+self.config["newline"]
                                                lines1.append(line1)
                                                if line1 not in nodesmap.keys():
                                                    nodesmap[line1]=nodes.cell(row=r,column=1).value
                                            if node_num!=None and int(node_num)==int(rdata["B"]):
                                                rdata["I"]=nodes.cell(row=r,column=2).value
                                                rdata["J"]=nodes.cell(row=r,column=3).value
                                                rdata["K"]=nodes.cell(row=r,column=4).value
                                                rdata["L"]=nodes.cell(row=r,column=5).value
                                                rdata["M"]=nodes.cell(row=r,column=6).value
                                                line2=self.config["split"].join([str(rdata["I"]),str(rdata["J"]),str(rdata["K"])])+self.config["split"]+"1"+self.config["split"]+"1.000000"+self.config["split"]+node_name+self.config["newline"]
                                                lines2.append(line2)
                                                if line2 not in nodesmap.keys():
                                                    nodesmap[line2]=nodes.cell(row=r,column=1).value
                                            r+=1
                                        if rdata["A"]<=rdata["B"]:
                                            sheet.append(rdata)
                                nodeslst=[]
                                for enode in lines1+lines2:
                                    if enode not in nodeslst:
                                        nodeslst.append(enode)
                                # 所有索引组合为1个列表并进行保留顺序去重
                                self.logger.info("正在为 %s 生成.node文件" %file)
                                with open(os.path.join(workdir,"Results",file+".node"),mode="w",encoding="utf-8") as writer:
                                    writer.writelines(nodeslst)
                                    writer.write(self.config["newline"])
                                self.logger.info("正在为 %s 生成.edge文件" %file)
                                node=numpy.zeros((len(nodeslst),len(nodeslst)),float)
                                if node.size==0:
                                    self.logger.warning("边缘矩阵大小为0")
                                else:
                                    self.logger.debug("边缘矩阵大小：(%d,%d)" %node.shape)
                                for idx,val in numpy.ndenumerate(node):
                                    if val!=0:
                                        self.logger.warning("边缘矩阵在索引(%d,%d)处存在非零初始化值 %f" %(idx[0],idx[1],val))
                                    node[idx]=matrix[nodesmap[nodeslst[idx[0]]]-1,nodesmap[nodeslst[idx[1]]]-1]
                                numpy.savetxt(os.path.join(workdir,"Results",file+".edge"),node,fmt="%s",delimiter=self.config["split"],newline=self.config["newline"],encoding="utf-8")
                                if self.config["selected_node_enabled"]==True:
                                    selected_nodes_list=numpy.loadtxt(selected_nodes,int)
                                    selected_result=numpy.zeros((len(selected_nodes_list),len(selected_nodes_list)),float)
                                    if selected_result.size==0:
                                        self.logger.warning("筛选结果矩阵大小为0")
                                    else:
                                        self.logger.debug("筛选结果矩阵大小为(%d,%d)" %selected_result.shape)
                                    for idx,val in numpy.ndenumerate(selected_result):
                                        if val!=0:
                                            self.logger.warning("筛选结果矩阵在索引(%d,%d)处存在非零初始化值 %f" %(idx[0],idx[1],val))
                                        selected_result[idx]=matrix[selected_nodes_list[idx[0]]-1,selected_nodes_list[idx[1]]-1]
                                    numpy.savetxt(os.path.join(workdir,"Results",file),selected_result,fmt="%s",delimiter=self.config["split"],newline=self.config["newline"],encoding="utf-8")
                                    self.logger.info("文件 %s 的筛选结果已保存" %file)
                                matrix[numpy.where(numpy.fabs(matrix)<target)]=0
                                # 将皮尔逊系数绝对值小于目标值的位置的皮尔逊系数值设为0
                                matrix=numpy.fabs(matrix)
                                # 将皮尔逊系数绝对值大于等于目标值的位置的皮尔逊系数值设为绝对值
                                datas.append(matrix)
                                fig=plt.figure()
                                ax=fig.add_subplot(111)
                                cax=ax.matshow(matrix,cmap=plt.cm.get_cmap(self.config["cmap"]))
                                fig.colorbar(cax)
                                fig.savefig(os.path.join(workdir,"Results",file+".png"))
                                plt.close(fig)
                            self.update_progress_signal.emit(int((files.index(file)+1)/len(file)*100))
                    results.save(os.path.join(workdir,"Results","maps.xlsx"))
                    results.close()
                    self.logger.info("正在生成 %d 张图像的公共部分" %(len(datas)))
                    self.busy_progress.emit()
                    common=self.get_common_part(datas,overnum)
                    max_num=0
                    if common[numpy.nonzero(common)].size!=0:
                        max_num=numpy.max(common[numpy.nonzero(common)])
                    else:
                        self.logger.warning("公共部分矩阵似乎不存在非零部分？")
                    # 取公共部分矩阵的非零最大值
                    max_pos=list()
                    for x,y in numpy.array(numpy.where(common==max_num)).T:
                        max_pos.append((x+1,y+1))
                    # 根据非零最大值求坐标列表
                    self.logger.debug("max_num=%d,max_pos=%s" %(max_num,max_pos))
                    fig=plt.figure()
                    ax=fig.add_subplot(111)
                    cax=ax.matshow(common,cmap=plt.cm.get_cmap(self.config["cmap"]))
                    fig.colorbar(cax)
                    fig.savefig(os.path.join(workdir,"Results","merged.png"))
                    plt.close(fig)
                    if max_num>0:
                        self.logger.info("处理过程中公共部分出现最多的点：%s，共出现 %d 次" %(max_pos,max_num))
                    else:
                        self.logger.info("未找到公共点")
                    if self.config["nii_enabled"]==True:
                        self.logger.info("正在生成nii文件中所有节点的值的列表")
                        niiimg=nibabel.load(self.config["nii_path"])
                        affine=niiimg.affine
                        niidata=numpy.asanyarray(niiimg.dataobj)
                        niiwb=openpyxl.Workbook()
                        niisheet=niiwb.active
                        niisheet.title="nii数据结果"
                        chat_head={"A1":"x","B1":"y","C1":"z","D1":"nii文件中的值"}
                        for key in chat_head.keys():
                            niisheet[key]=chat_head[key]
                            niisheet[key].alignment=center_style
                        r=self.config["node_start"]
                        for row in nodes.iter_rows(min_row=self.config["node_start"]):
                            pos=[cell.value for cell in list(row)[1:4]]
                            if any([p==None for p in pos]):
                                break
                            vpos=apply_affine(numpy.linalg.inv(affine),pos)
                            val=niidata[round(vpos[0]),round(vpos[1]),round(vpos[2])]
                            chatdata=pos
                            chatdata.extend([val])
                            niisheet.append(chatdata)
                            r+=1
                        niiwb.save(os.path.join(workdir,"Results","nii_values.xlsx"))
                    if self.config["dicom_enabled"]==True:
                        self.logger.info("正在从DICOM文件夹中提取患者信息")
                        wb=openpyxl.Workbook()
                        fr=wb.active
                        fr.title="FunRawResult"
                        tr=wb.create_sheet("T1RawResult")
                        # 最后去重的结果按文件夹放到这两张表
                        fs=wb.create_sheet("FunRaw")
                        ts=wb.create_sheet("T1Raw")
                        # 每个文件夹内的每一张图像获取到的原始数据
                        fs.append(["文件路径","患者ID","患者姓名","患者生日","患者年龄","患者性别","患者身高","患者体重"])
                        ts.append(["文件路径","患者ID","患者姓名","患者生日","患者年龄","患者性别","患者身高","患者体重"])
                        fr.append(["患者ID","患者姓名","患者生日","患者年龄","患者性别","患者身高","患者体重"])
                        tr.append(["患者ID","患者姓名","患者生日","患者年龄","患者性别","患者身高","患者体重"])
                        ffailed=0
                        fold=[]
                        tfailed=0
                        told=[]
                        for root,dirs,fnames in os.walk(os.path.join(self.config["dicom_path"],"FunRaw")):
                            for fname in fnames:
                                try:
                                    ds=dcmread(os.path.join(root,fname))
                                except:
                                    self.logger.error("文件 %s 读取失败，将跳过" %os.path.join(root,fname))
                                    ffailed+=1
                                else:
                                    self.logger.debug("加载文件 %s 成功" %os.path.join(root,fname))
                                    data,datab=self.handle_ds(ds)
                                    if datab!=fold:
                                        fold=datab
                                        fr.append(datab)
                                    fs.append(data)
                        for root,dirs,fnames in os.walk(os.path.join(self.config["dicom_path"],"T1Raw")):
                            for fname in fnames:
                                try:
                                    ds=dcmread(os.path.join(root,fname))
                                except:
                                    self.logger.error("文件 %s 读取失败，将跳过" %os.path.join(root,fname))
                                    tfailed+=1
                                else:
                                    self.logger.debug("加载文件 %s 成功" %os.path.join(root,fname))
                                    data,datab=self.handle_ds(ds)
                                    if datab!=told:
                                        told=datab
                                        tr.append(datab)
                                    ts.append(data)
                        wb.save(os.path.join(workdir,"Results","info.xlsx"))
                        self.logger.info("信息提取完成，FunRaw文件夹失败文件数：%d，T1Raw文件夹失败文件数：%d" %(ffailed,tfailed))
                    mins,secs=divmod(time.time()-start_time,60)
                    hrs,mins=divmod(mins,60)
                    self.logger.info("所有任务执行完成，共计用时 %02d:%02d:%02d" %(hrs,mins,secs))
                    self.finish_signal.emit(0)
                elif target>1 or target<0:
                    self.logger.error("皮尔逊系数为一个[0,1]的浮点数")
                    self.finish_signal.emit(1)
                elif overnum<=0 or overnum>len(files):
                    self.logger.error("输入的图片张数非法，应属于(0,%d]" %len(files))
                    self.finish_signal.emit(1)
                else:
                    self.logger.error("未知错误")
                    self.finish_signal.emit(1)
class GUIHandler(logging.Handler):
    def __init__(self,update_log_signal:pyqtBoundSignal):
        super().__init__()
        self.update_log_signal=update_log_signal
    def emit(self,record:logging.LogRecord):
        self.update_log_signal.emit(self.format(record))
class SettingDialog(QDialog):
    update_config_signal=pyqtSignal(dict)
    def __init__(self,parent:QWidget):
        super().__init__()
        self.m_flag=False
        self.logger=logging.getLogger()
        self.parent_=parent
        with open("config.json","r",encoding="utf-8") as reader:
            self.config=json.load(reader)
        layout=QGridLayout()
        layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.setLayout(layout)
        self.setWindowFlags(Qt.WindowType.FramelessWindowHint|Qt.WindowType.Tool)
        self.setAutoFillBackground(True)
        self.setWindowOpacity(parent.windowOpacity())
        title=QLabel("设置")
        title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        title.setStyleSheet("QLabel{border:none;border-radius:5px;background:transparent;color:#9AD3BC;font-size:20px;}")
        layout.addWidget(title,0,1)
        close_btn=QPushButton()
        close_btn.setStyleSheet("QPushButton{background:#FFE3ED;border-radius:5px;border:none;}QPushButton:hover{background:#EC524B;}")
        close_btn.setToolTip("保存设置并退出")
        close_btn.clicked.connect(self.close)
        layout.addWidget(close_btn,0,0)
        content=QGridLayout()
        self.debug=QCheckBox("调试模式")
        self.debug.setChecked(bool(self.config["debug"]))
        self.debug.setStyleSheet("QCheckBox::indicator{width:10px;height:10px;border:none;border-radius:5px;background:#9BE3DE;}QCheckBox::indicator:unchecked{background:#BEEBE9;}QCheckBox::indicator:unchecked:hover{background:#9AD3BC;}QCheckBox::indicator:checked{background:#95E1D3;}QCheckBox::indicator:checked:hover{background:#98DED9;}")
        content.addWidget(self.debug,1,0)
        self.minimal_to_tray=QCheckBox("最小化到托盘")
        self.minimal_to_tray.setChecked(self.config["minimal_to_tray"])
        self.minimal_to_tray.setStyleSheet("QCheckBox::indicator{width:10px;height:10px;border:none;border-radius:5px;background:#9BE3DE;}QCheckBox::indicator:unchecked{background:#BEEBE9;}QCheckBox::indicator:unchecked:hover{background:#9AD3BC;}QCheckBox::indicator:checked{background:#95E1D3;}QCheckBox::indicator:checked:hover{background:#98DED9;}")
        content.addWidget(self.minimal_to_tray,2,0)
        process_info=QHBoxLayout()
        process_info_label=QLabel("处理配置：")
        process_info_label.setStyleSheet("QLabel{background:transparent;border:none;}")
        process_info.addWidget(process_info_label,1)
        self.process_info=QLineEdit()
        self.process_info.setText(str(self.config["process_info"]))
        self.process_info.setStyleSheet("QLineEdit{border:1px solid #F3EAC2;border-radius:5px;background:transparent;}QLineEdit:hover{border:1px solid #F5B461;}")
        process_info.addWidget(self.process_info,8)
        process_info_btn=QPushButton("浏览")
        process_info_btn.setFixedHeight(20)
        process_info_btn.setStyleSheet("QPushButton{background:#BEEBE9;border-radius:5px;border:none;}QPushButton:hover{background:#F3EAC2;}")
        process_info_btn.clicked.connect(self.browse_process_info)
        process_info.addWidget(process_info_btn,1)
        content.addLayout(process_info,0,0)
        layout.addLayout(content,1,1)
    def close(self):
        config={
            "debug":self.debug.isChecked(),
            "process_info":self.process_info.text(),
            "current_pos":[
                self.parent_.pos().x()|0,
                self.parent_.pos().y()|0
            ],
            "minimal_to_tray":self.minimal_to_tray.isChecked()
        }
        with open("config.json","w",encoding="utf-8") as writer:
            writer.write(json.dumps(config,sort_keys=True,indent=4,ensure_ascii=False))
            self.logger.debug("已保存配置")
        self.update_config_signal.emit(config)
        super().close()
    def browse_process_info(self):
        file,_=QFileDialog.getOpenFileName(caption="选择处理信息文件",directory=os.path.split(os.path.realpath(__file__))[0],filter="处理信息(*.json)")
        if file!="":
            self.process_info.setText(file)
        self.logger.debug("获取到的文件信息：%s" %file)
    def mousePressEvent(self, event:QMouseEvent):
        self.logger.debug("触发鼠标按压事件")
        super().mousePressEvent(event)
        self.setFocus()
        self.m_flag=True
        if event.button()==Qt.MouseButton.LeftButton and self.isMaximized()==False and self.hasFocus()==True:
            self.old_pos=event.globalPosition() #获取鼠标相对窗口的位置
            self.logger.debug("已获取鼠标位置")
    def mouseMoveEvent(self, event:QMouseEvent):
        self.logger.debug("触发鼠标移动事件")
        super().mouseMoveEvent(event)
        if self.m_flag==True:
            delta_x=int(event.globalPosition().x()-self.old_pos.x())
            delta_y=int(event.globalPosition().y()-self.old_pos.y())
            self.move(self.x()+delta_x,self.y()+delta_y)#更改窗口位置
            self.logger.debug("已更改窗口位置")
            self.old_pos=event.globalPosition()
            self.setCursor(QCursor(Qt.CursorShape.SizeAllCursor))  #更改鼠标图标
    def mouseReleaseEvent(self, event:QMouseEvent):
        self.logger.debug("触发鼠标释放事件")
        super().mouseReleaseEvent(event)
        self.m_flag=False
        self.setCursor(QCursor(Qt.CursorShape.ArrowCursor))
class ProcessInfoEditor(QDialog):
    update_config_signal=pyqtSignal(dict)
    def __init__(self,parent:QWidget):
        super().__init__()
        self.m_flag=False
        self.logger=logging.getLogger()
        layout=QGridLayout()
        self.setLayout(layout)
        self.setWindowFlags(Qt.WindowType.FramelessWindowHint|Qt.WindowType.Tool)
        self.setAutoFillBackground(True)
        self.setWindowOpacity(parent.windowOpacity())
        close_btn=QPushButton()
        close_btn.setToolTip("关闭")
        close_btn.setStyleSheet("QPushButton{background:#FFE3ED;border-radius:5px;border:none;}QPushButton:hover{background:#EC524B;}")
        close_btn.clicked.connect(self.close)
        layout.addWidget(close_btn,0,0)
        title=QLabel("处理信息文件编辑器")
        title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        title.setStyleSheet("QLabel{border:none;border-radius:5px;background:transparent;color:#9AD3BC;font-size:20px;}")
        layout.addWidget(title,0,1)
        content=QGridLayout()
        open_exists=QHBoxLayout()
        open_exists_label=QLabel("打开现有文件：")
        open_exists_label.setStyleSheet("QLabel{background:transparent;border:none;}")
        open_exists.addWidget(open_exists_label)
        self.open_exists_edit=QLineEdit()
        self.open_exists_edit.setStyleSheet("QLineEdit{border:1px solid #F3EAC2;border-radius:5px;background:transparent;}QLineEdit:hover{border:1px solid #F5B461;}")
        open_exists.addWidget(self.open_exists_edit)
        open_exists_btn=QPushButton("浏览")
        open_exists_btn.setToolTip("浏览文件")
        open_exists_btn.setFixedSize(40,20)
        open_exists_btn.setStyleSheet("QPushButton{background:#9BE3DE;border:none;border-radius:5px}QPushButton:hover{background:#9AD3BC;}")
        open_exists_btn.clicked.connect(self.open_exists)
        open_exists.addWidget(open_exists_btn)
        content.addLayout(open_exists,0,0)
        add_label=QHBoxLayout()
        add_label_label=QLabel("添加标签：")
        add_label_label.setStyleSheet("QLabel{background:transparent;border:none;}")
        add_label.addWidget(add_label_label)
        self.add_label_combo=QComboBox()
        combostyle='''
            QComboBox{border-radius:5px;border:none;background:transparent;}
            QComboBox:hover{border:1px solid #F5B461;}
            QComboBox::drop-down{background:transparent;}
            QComboBox::drop-down:hover{background:#F5B461;}
            QComboBox::down-arrow{width:10px;height:10px;background:#9BE3DE;border-radius:5px;}
            QComboBox::down-arrow:on{background:#9AD3BC;}
            QComboBox QAbstractItemView{outline:none;border:none;border-radius:5px;}
            QComboBox QAbstractItemView:hover{border:1px solid #F5B461;}
            QComboBox QAbstractItemView::item{background:#F3EAC2;}
            QComboBox QAbstractItemView::item:selected{background:#F5B461;color:#F3EAC2;}
            '''
        self.add_label_combo_data={"中文":"C","英文":"E","无":"N"}
        for key,value in self.add_label_combo_data.items():
            self.add_label_combo.addItem(key,value)
        self.add_label_combo.setToolTip("添加标签到.node文件，需要节点文件内拥有相应数据")
        self.add_label_combo.setStyleSheet(combostyle)
        self.add_label_combo.setCurrentText("无")
        add_label.addWidget(self.add_label_combo)
        content.addLayout(add_label,1,0)
        cmap=QHBoxLayout()
        cmap_label=QLabel("CMAP：")
        cmap_label.setStyleSheet("QLabel{background:transparent;border:none;}")
        cmap.addWidget(cmap_label)
        self.cmap_edit=QLineEdit("viridis")
        self.cmap_edit.setToolTip("matplotlib绘图的颜色\n参考 https://matplotlib.org/stable/tutorials/colors/colormaps.html 进行自定义")
        self.cmap_edit.setStyleSheet("QLineEdit{border:1px solid #F3EAC2;border-radius:5px;background:transparent;}QLineEdit:hover{border:1px solid #F5B461;}")
        cmap.addWidget(self.cmap_edit)
        content.addLayout(cmap,1,1)
        newline=QHBoxLayout()
        newline_label=QLabel("换行符：")
        newline_label.setStyleSheet("QLabel{background:transparent;border:none;}")
        newline.addWidget(newline_label)
        self.newline_edit=QLineEdit("\\n")
        self.newline_edit.setToolTip("生成的矩阵文本文件使用的换行符")
        self.newline_edit.setStyleSheet("QLineEdit{border:1px solid #F3EAC2;border-radius:5px;background:transparent;}QLineEdit:hover{border:1px solid #F5B461;}")
        newline.addWidget(self.newline_edit)
        content.addLayout(newline,2,0)
        split=QHBoxLayout()
        split_label=QLabel("分隔符：")
        split_label.setStyleSheet("QLabel{background:transparent;border:none;}")
        split.addWidget(split_label)
        self.split_edit=QLineEdit("\\t")
        self.split_edit.setToolTip("生成的矩阵文本文件使用的用于分隔单行中各个数字的分隔符")
        self.split_edit.setStyleSheet("QLineEdit{border:1px solid #F3EAC2;border-radius:5px;background:transparent;}QLineEdit:hover{border:1px solid #F5B461;}")
        split.addWidget(self.split_edit)
        content.addLayout(split,2,1)
        node_start=QHBoxLayout()
        node_start_label=QLabel("节点起始行：")
        node_start_label.setStyleSheet("QLabel{background:transparent;border:none;}")
        node_start.addWidget(node_start_label)
        self.node_start_edit=QLineEdit("3")
        self.node_start_edit.setToolTip("节点文件的起始行数，用于跳过表头的影响，从1开始")
        self.node_start_edit.setValidator(QRegularExpressionValidator(QRegularExpression("[1-9][0-9]{0,}")))
        self.node_start_edit.setStyleSheet("QLineEdit{border:1px solid #F3EAC2;border-radius:5px;background:transparent;}QLineEdit:hover{border:1px solid #F5B461;}")
        node_start.addWidget(self.node_start_edit)
        content.addLayout(node_start,3,0)
        workdir=QHBoxLayout()
        workdir_label=QLabel("工作目录：")
        workdir_label.setStyleSheet("QLabel{background:transparent;border:none;}")
        workdir.addWidget(workdir_label)
        self.workdir_edit=QLineEdit()
        self.workdir_edit.setToolTip("存放矩阵文件的文件夹")
        self.workdir_edit.setStyleSheet("QLineEdit{border:1px solid #F3EAC2;border-radius:5px;background:transparent;}QLineEdit:hover{border:1px solid #F5B461;}")
        workdir.addWidget(self.workdir_edit)
        workdir_btn=QPushButton("浏览")
        workdir_btn.setToolTip("浏览文件")
        workdir_btn.setFixedSize(40,20)
        workdir_btn.setStyleSheet("QPushButton{background:#9BE3DE;border:none;border-radius:5px}QPushButton:hover{background:#9AD3BC;}")
        workdir_btn.clicked.connect(self.browse_workdir)
        workdir.addWidget(workdir_btn)
        content.addLayout(workdir,4,0)
        nodepath=QHBoxLayout()
        nodepath_label=QLabel("节点信息文件：")
        nodepath_label.setStyleSheet("QLabel{background:transparent;border:none;}")
        nodepath.addWidget(nodepath_label)
        self.nodepath_edit=QLineEdit()
        self.nodepath_edit.setToolTip("选择节点文件的位置\n节点文件格式：\n节点序号 x y z 英文名 中文名")
        self.nodepath_edit.setStyleSheet("QLineEdit{border:1px solid #F3EAC2;border-radius:5px;background:transparent;}QLineEdit:hover{border:1px solid #F5B461;}")
        nodepath.addWidget(self.nodepath_edit)
        nodepath_btn=QPushButton("浏览")
        nodepath_btn.setToolTip("浏览文件")
        nodepath_btn.setFixedSize(40,20)
        nodepath_btn.setStyleSheet("QPushButton{background:#9BE3DE;border:none;border-radius:5px}QPushButton:hover{background:#9AD3BC;}")
        nodepath_btn.clicked.connect(self.browse_nodepath)
        nodepath.addWidget(nodepath_btn)
        content.addLayout(nodepath,5,0)
        target=QHBoxLayout()
        target_label=QLabel("目标皮尔逊绝对值：")
        target_label.setStyleSheet("QLabel{background:transparent;border:none;}")
        target.addWidget(target_label)
        self.target_edit=QLineEdit("0.6")
        self.target_edit.setToolTip("目标皮尔逊相关系数的绝对值")
        self.target_edit.setStyleSheet("QLineEdit{border:1px solid #F3EAC2;border-radius:5px;background:transparent;}QLineEdit:hover{border:1px solid #F5B461;}")
        self.target_edit.setValidator(QDoubleValidator(0.0,1.0,0))
        target.addWidget(self.target_edit)
        content.addLayout(target,6,0)
        overnum=QHBoxLayout()
        overnum_label=QLabel("最少图片数：")
        overnum_label.setStyleSheet("QLabel{background:transparent;border:none;}")
        overnum.addWidget(overnum_label)
        self.overnum_edit=QLineEdit("10")
        self.overnum_edit.setToolTip("可以被计入公共部分所需要的满足绝对值不小于目标皮尔逊绝对值的最少图片张数")
        self.overnum_edit.setStyleSheet("QLineEdit{border:1px solid #F3EAC2;border-radius:5px;background:transparent;}QLineEdit:hover{border:1px solid #F5B461;}")
        self.overnum_edit.setValidator(QRegularExpressionValidator(QRegularExpression("[1-9][0-9]{0,}")))
        overnum.addWidget(self.overnum_edit)
        content.addLayout(overnum,6,1)
        selected_node_path=QHBoxLayout()
        selected_node_path_label=QLabel("筛选的节点列表")
        selected_node_path_label.setStyleSheet("QLabel{background:transparent;border:none;}")
        selected_node_path.addWidget(selected_node_path_label)
        self.selected_node_path_edit=QLineEdit()
        self.selected_node_path_edit.setToolTip("用于筛选符合要求的矩阵的节点列表，格式为1行1个节点数，需要提供节点文件\n留空则禁用筛选功能")
        self.selected_node_path_edit.setStyleSheet("QLineEdit{border:1px solid #F3EAC2;border-radius:5px;background:transparent;}QLineEdit:hover{border:1px solid #F5B461;}")
        selected_node_path.addWidget(self.selected_node_path_edit)
        selected_node_path_btn=QPushButton("浏览")
        selected_node_path_btn.setToolTip("浏览文件")
        selected_node_path_btn.setFixedSize(40,20)
        selected_node_path_btn.setStyleSheet("QPushButton{background:#9BE3DE;border:none;border-radius:5px}QPushButton:hover{background:#9AD3BC;}")
        selected_node_path_btn.clicked.connect(self.browse_selected_node)
        selected_node_path.addWidget(selected_node_path_btn)
        content.addLayout(selected_node_path,7,0)
        nii_path=QHBoxLayout()
        nii_path_label=QLabel(".nii文件位置：")
        nii_path_label.setStyleSheet("QLabel{background:transparent;border:none;}")
        nii_path.addWidget(nii_path_label)
        self.nii_path_edit=QLineEdit()
        self.nii_path_edit.setToolTip("输出所有节点的值所需要的.nii文件的位置，留空则禁用此功能")
        self.nii_path_edit.setStyleSheet("QLineEdit{border:1px solid #F3EAC2;border-radius:5px;background:transparent;}QLineEdit:hover{border:1px solid #F5B461;}")
        nii_path.addWidget(self.nii_path_edit)
        nii_path_btn=QPushButton("浏览")
        nii_path_btn.setToolTip("浏览文件")
        nii_path_btn.setFixedSize(40,20)
        nii_path_btn.setStyleSheet("QPushButton{background:#9BE3DE;border:none;border-radius:5px}QPushButton:hover{background:#9AD3BC;}")
        nii_path_btn.clicked.connect(self.browse_nii)
        nii_path.addWidget(nii_path_btn)
        content.addLayout(nii_path,8,0)
        dicom=QHBoxLayout()
        dicom_label=QLabel("DICOM文件夹：")
        dicom_label.setStyleSheet("QLabel{background:transparent;border:none;}")
        dicom.addWidget(dicom_label)
        self.dicom_edit=QLineEdit()
        self.dicom_edit.setToolTip("用于提取用户信息的DICOM文件夹的位置，即含有FunRaw和T1Raw的那个文件夹，留空则禁用提取用户信息的功能")
        self.dicom_edit.setStyleSheet("QLineEdit{border:1px solid #F3EAC2;border-radius:5px;background:transparent;}QLineEdit:hover{border:1px solid #F5B461;}")
        dicom.addWidget(self.dicom_edit)
        dicom_btn=QPushButton("浏览")
        dicom_btn.setToolTip("浏览文件")
        dicom_btn.setFixedSize(40,20)
        dicom_btn.setStyleSheet("QPushButton{background:#9BE3DE;border:none;border-radius:5px}QPushButton:hover{background:#9AD3BC;}")
        dicom_btn.clicked.connect(self.browse_dicom)
        dicom.addWidget(dicom_btn)
        content.addLayout(dicom,9,0)
        self.apply=QCheckBox("保存并应用")
        self.apply.setToolTip("保存的同时应用这个处理信息文件到脚本配置")
        self.apply.setStyleSheet("QCheckBox::indicator{width:10px;height:10px;border:none;border-radius:5px;background:#9BE3DE;}QCheckBox::indicator:unchecked{background:#BEEBE9;}QCheckBox::indicator:unchecked:hover{background:#9AD3BC;}QCheckBox::indicator:checked{background:#95E1D3;}QCheckBox::indicator:checked:hover{background:#98DED9;}")
        content.addWidget(self.apply,9,1)
        save=QPushButton("保存(&S)")
        save.setToolTip("保存当前配置")
        save.setFixedHeight(20)
        save.setStyleSheet("QPushButton{background:#9BE3DE;border:none;border-radius:5px}QPushButton:hover{background:#9AD3BC;}")
        save.clicked.connect(self.save)
        layout.addLayout(content,1,1)
        layout.addWidget(save,2,1)
    def open_exists(self):
        file,_=QFileDialog.getOpenFileName(caption="选择处理信息文件",directory=os.path.split(os.path.realpath(__file__))[0],filter="处理信息(*.json)")
        self.logger.debug("获取到的文件信息：%s" %file)
        if file!="":
            try:
                with open(file,"r",encoding="utf-8") as reader:
                    conf=json.load(reader)
            except:
                self.logger.error("读取文件失败")
            else:
                try:
                    idx=list(self.add_label_combo_data.values()).index(conf["add_label"])
                    self.add_label_combo.setCurrentIndex(idx)
                    self.cmap_edit.setText(conf["cmap"])
                    self.newline_edit.setText(conf["newline"])
                    self.split_edit.setText(conf["split"])
                    self.node_start_edit.setText(str(conf["node_start"]))
                    self.workdir_edit.setText(conf["workdir"])
                    self.nodepath_edit.setText(conf["nodepath"])
                    self.target_edit.setText(str(conf["target"]))
                    self.overnum_edit.setText(str(conf["overnum"]))
                    self.selected_node_path_edit.setText(conf["selected_nodepath"])
                    self.nii_path_edit.setText(conf["nii_path"])
                    self.dicom_edit.setText(conf["dicom_path"])
                except Exception as e:
                    self.logger.debug("出错原因：%s" %e)
                    self.logger.warning("文件中存在错误，我们已经将错误部分恢复为默认值")
                self.open_exists_edit.setText(file)
    def browse_workdir(self):
        file=QFileDialog.getExistingDirectory(caption="选择工作文件夹",directory=os.path.split(os.path.realpath(__file__))[0])
        self.workdir_edit.setText(file)
        self.logger.debug("获取到的文件夹信息：%s" %file)
    def browse_nodepath(self):
        file,_=QFileDialog.getOpenFileName(caption="选择节点文件",directory=os.path.split(os.path.realpath(__file__))[0],filter="xlsx格式节点信息(*.xlsx);;csv格式节点信息(*.csv *.txt)")
        self.nodepath_edit.setText(file)
        self.logger.debug("获取到的文件信息：%s" %file)
    def browse_selected_node(self):
        file,_=QFileDialog.getOpenFileName(caption="选择节点列表",directory=os.path.split(os.path.realpath(__file__))[0],filter="csv格式节点信息(*.csv *.txt)")
        self.selected_node_path_edit.setText(file)
        self.logger.debug("获取到的文件信息：%s" %file)
    def browse_nii(self):
        file,_=QFileDialog.getOpenFileName(caption="选择.nii数据文件",directory=os.path.split(os.path.realpath(__file__))[0],filter="数据文件(*.nii *.nii.gz)")
        self.nii_path_edit.setText(file)
        self.logger.debug("获取到的文件信息：%s" %file)
    def browse_dicom(self):
        file=QFileDialog.getExistingDirectory(caption="选择DICOM文件夹",directory=os.path.split(os.path.realpath(__file__))[0])
        if os.path.isdir(os.path.join(file,"FunRaw")) and os.path.isdir(os.path.join(file,"T1Raw")):
            self.dicom_edit.setText(file)
            self.logger.debug("获取到的文件夹信息：%s" %file)
        else:
            QMessageBox.critical(self,"错误","DICOM文件夹需要拥有FunRaw和T1Raw文件夹")        
    def save(self):
        conf={
            "add_label":self.add_label_combo.currentData(),
            "cmap":self.cmap_edit.text(),
            "newline":self.newline_edit.text(),
            "split":self.split_edit.text(),
            "node_start":int(self.node_start_edit.text()),
            "workdir":self.workdir_edit.text(),
            "nodepath":self.nodepath_edit.text(),
            "target":float(self.target_edit.text()),
            "overnum":int(self.overnum_edit.text()),
            "selected_nodepath":self.selected_node_path_edit.text(),
            "selected_node_enabled":bool(self.selected_node_path_edit.text()),
            "nii_path":self.nii_path_edit.text(),
            "nii_enabled":bool(self.nii_path_edit.text()),
            "dicom_path":self.dicom_edit.text(),
            "dicom_enabled":bool(self.dicom_edit.text())
        }
        if self.open_exists_edit.text()!="" and os.path.isfile(self.open_exists_edit.text()):
            path=self.open_exists_edit.text()
        else:
            path,_=QFileDialog.getSaveFileName(caption="保存至",directory=os.path.split(os.path.realpath(__file__))[0],filter="处理信息文件(*.json)")
        if path!="":
            with open(path,"w",encoding="utf-8") as writer:
                writer.write(json.dumps(conf,indent=4,sort_keys=True,ensure_ascii=False))
            self.logger.info("已保存至 %s" %path)
            if self.open_exists_edit.text()=="":
                self.open_exists_edit.setText(path)
            if self.apply.isChecked()==True:
                self.update_config_signal.emit({"process_info":path})
    def mousePressEvent(self, event:QMouseEvent):
        self.logger.debug("触发鼠标按压事件")
        super().mousePressEvent(event)
        self.setFocus()
        self.m_flag=True
        if event.button()==Qt.MouseButton.LeftButton and self.isMaximized()==False and self.hasFocus()==True:
            self.old_pos=event.globalPosition() #获取鼠标相对窗口的位置
            self.logger.debug("已获取鼠标位置")
    def mouseMoveEvent(self, event:QMouseEvent):
        self.logger.debug("触发鼠标移动事件")
        super().mouseMoveEvent(event)
        if self.m_flag==True:
            delta_x=int(event.globalPosition().x()-self.old_pos.x())
            delta_y=int(event.globalPosition().y()-self.old_pos.y())
            self.move(self.x()+delta_x,self.y()+delta_y)#更改窗口位置
            self.logger.debug("已更改窗口位置")
            self.old_pos=event.globalPosition()
            self.setCursor(QCursor(Qt.CursorShape.SizeAllCursor))  #更改鼠标图标
    def mouseReleaseEvent(self, event:QMouseEvent):
        self.logger.debug("触发鼠标释放事件")
        super().mouseReleaseEvent(event)
        self.m_flag=False
        self.setCursor(QCursor(Qt.CursorShape.ArrowCursor))
class UI(QMainWindow):
    update_log_signal=pyqtSignal(str)
    update_progress_signal=pyqtSignal(int)
    busy_progress_signal=pyqtSignal()
    finish_signal=pyqtSignal(int)
    def __init__(self):
        super().__init__()
        if os.path.exists("config.json")==False:
            self.create_config()
        with open("config.json","r",encoding="utf-8") as reader:
            self.conf=json.load(reader)
        self.m_flag=False
        self.need_tip=True
        icon=QIcon()
        pixmap=QPixmap()
        pixmap.loadFromData(base64.b64decode(ICON))
        icon.addPixmap(pixmap)
        self.setWindowIcon(icon)
        self.setWindowFlag(Qt.WindowType.FramelessWindowHint)
        self.setWindowOpacity(0.9)
        self.setAttribute(Qt.WidgetAttribute.WA_TranslucentBackground)
        self.setAutoFillBackground(True)
        self.resize(1024,768)
        self.setWindowTitle(APPID)
        self.logger=logging.getLogger()
        self.logger.addHandler(GUIHandler(update_log_signal=self.update_log_signal))
        self.logger.addHandler(logging.FileHandler(filename=__file__+".log",mode="w",encoding="utf-8"))
        self.logger.setLevel(logging.INFO)
        for handler in self.logger.handlers:
            handler.setLevel(logging.INFO)
            handler.setFormatter(logging.Formatter(fmt="%(asctime)s-%(levelname)s-%(message)s",datefmt="%Y-%m-%d %H:%M:%S"))
        self.move(self.conf["current_pos"][0],self.conf["current_pos"][1])
        if self.conf["debug"]==True:
            self.logger.setLevel(logging.DEBUG)
            for handler in self.logger.handlers:
                handler.setLevel(logging.DEBUG)
        cw=QWidget()
        cl=QGridLayout()
        cw.setLayout(cl)
        self.setCentralWidget(cw)
        control=QVBoxLayout()
        self.min_btn=QPushButton()
        self.min_btn.setToolTip("最小化")
        self.min_btn.setStyleSheet("QPushButton{background:#BEEBE9;border-radius:5px;border:none;}QPushButton:hover{background:#F3EAC2;}")
        self.min_btn.clicked.connect(self.showMinimized)
        control.addWidget(self.min_btn)
        self.close_btn=QPushButton()
        self.close_btn.setToolTip("关闭")
        self.close_btn.setStyleSheet("QPushButton{background:#FFE3ED;border-radius:5px;border:none;}QPushButton:hover{background:#EC524B;}")
        self.close_btn.clicked.connect(self.close)
        control.addWidget(self.close_btn)
        control.setAlignment(Qt.AlignmentFlag.AlignCenter)
        cl.addLayout(control,0,0)
        title=QLabel(APPID)
        title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        title.setStyleSheet("QLabel{border:none;border-radius:5px;background:transparent;color:#9AD3BC;font-size:60px;}")
        cl.addWidget(title,0,1)
        self.displayer=QPlainTextEdit()
        self.displayer.setReadOnly(True)
        self.displayer.setStyleSheet("QPlainTextEdit{background:#F3EAC2;border:none;border-radius:5px;}QScrollBar:vertical,QScrollBar::handle:vertical{background:#F3EAC2;border:none;border-radius:8px;width:16px;}QScrollBar::handle:vertical:hover{background:#F5B461;}QScrollBar::add-page:vertical,QScrollBar::sub-page:vertical{background:#FFFDF9;border:none;border-radius:8px;width:16px;}QScrollBar::down-arrow:vertical,QScrollBar::up-arrow:vertical{background:#F5B461;border:none;border-radius:8px;width:16px;height:16px;}QScrollBar::sub-line:vertical,QScrollBar::add-line:vertical{background:transparent;border:none;}")
        cl.addWidget(self.displayer,1,1)
        process=QHBoxLayout()
        self.start_btn=QPushButton("开始(&S)")
        self.start_btn.setToolTip("开始处理")
        self.start_btn.setFixedHeight(40)
        self.start_btn.setDefault(True)
        self.start_btn.setStyleSheet("QPushButton{background:#9BE3DE;border:none;border-radius:5px}QPushButton:hover{background:#9AD3BC;}")
        self.start_btn.clicked.connect(self.start_process)
        cl.addWidget(self.start_btn,3,1)
        self.setting_btn=QPushButton("设置(&C)")
        self.setting_btn.setToolTip("设置")
        self.setting_btn.setFixedHeight(20)
        self.setting_btn.setStyleSheet("QPushButton{background:#9BE3DE;border:none;border-radius:5px}QPushButton:hover{background:#9AD3BC;}")
        self.setting_btn.clicked.connect(self.show_setting)
        process.addWidget(self.setting_btn)
        self.process_info_btn=QPushButton("处理信息编辑器(&P)")
        self.process_info_btn.setToolTip("编辑处理信息文件")
        self.process_info_btn.setFixedHeight(20)
        self.process_info_btn.setStyleSheet("QPushButton{background:#9BE3DE;border:none;border-radius:5px}QPushButton:hover{background:#9AD3BC;}")
        self.process_info_btn.clicked.connect(self.show_process_info_editor)
        process.addWidget(self.process_info_btn)
        process.setAlignment(Qt.AlignmentFlag.AlignCenter)
        cl.addLayout(process,4,1)
        self.progress=QProgressBar()
        self.progress.setValue(0)
        self.progress.setFormat("当前进度：%p%")
        self.progress.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.progress.setStyleSheet("QProgressBar{background:#F3EAC2;border:none;border-radius:5px;}QProgressBar::chunk{background:qlineargradient(x1:0,x2:1,y1:0,y2:0,stop:0 #F3EAC2,stop:1 #98E3DE);border:none;border-radius:5px;}")
        cl.addWidget(self.progress,2,1)
        self.tray=QSystemTrayIcon()
        self.tray.setIcon(icon)
        self.tray.setToolTip(APPID)
        self.tray.activated.connect(self.tray_activated)
        self.tray.show()
        self.finish_signal.connect(self.finish_process)
        self.update_progress_signal.connect(self.upgrade_progress)
        self.busy_progress_signal.connect(self.busy_progress)
        self.update_log_signal.connect(self.update_gui_log)
        self.logger.debug("主界面初始化完成")
        if self.conf["process_info"]!="":
            self.logger.info("当前设置的处理信息文件：%s" %self.conf["process_info"])
        else:
            self.logger.warning("当前没有设置处理信息文件")
    def create_config(self):
        default_config={
            "debug":False,
            "minimal_to_tray":False,
            "process_info":"",
            "current_pos":[0,0]
        }
        with open("config.json","w",encoding="utf-8") as writer:
            writer.write(json.dumps(default_config,sort_keys=True,indent=4,ensure_ascii=False))
    @pyqtSlot(dict)
    def update_config(self,config:dict):
        self.conf.update(config)
        with open("config.json","w",encoding="utf-8") as writer:
            writer.write(json.dumps(self.conf,sort_keys=True,ensure_ascii=False,indent=4))
        self.logger.debug("已更新配置文件并保存至磁盘")
    @pyqtSlot(QSystemTrayIcon.ActivationReason)
    def tray_activated(self,reason:QSystemTrayIcon.ActivationReason):
        if reason==QSystemTrayIcon.ActivationReason.DoubleClick:
            if self.isVisible()==True:
                self.setVisible(False)
                self.tray.setVisible(True)
            else:
                self.setVisible(True)
                self.setFocus()
    @pyqtSlot(str)
    def update_gui_log(self,log:str):
        self.displayer.appendPlainText(log)
        if self.displayer.textCursor().atEnd()==False:
            scrollbar=self.displayer.verticalScrollBar()
            scrollbar.setSliderPosition(scrollbar.maximum())
    def start_process(self):
        if self.conf["process_info"]!="":
            self.work_thread=QThread()
            self.p=Processor(infopath=self.conf["process_info"],
                    update_progress_signal=self.update_progress_signal,busy_progress=self.busy_progress_signal,finish_signal=self.finish_signal)
            self.p.moveToThread(self.work_thread)
            self.work_thread.started.connect(self.p.main)
            self.work_thread.start()
            self.logger.debug("已启动子线程")
            self.start_btn.setEnabled(False)
            self.setting_btn.setEnabled(False)
        else:
            self.logger.error("处理信息文件路径为空")
    @pyqtSlot(int)
    def finish_process(self,code:int):
        self.work_thread.quit()
        self.work_thread.wait()
        self.start_btn.setEnabled(True)
        self.setting_btn.setEnabled(True)
        self.progress.setRange(0,100)
        self.progress.setValue(0)
        self.logger.debug("已结束子线程")
        self.tray.setVisible(True)
        if code==0:
            self.tray.showMessage("fMRI处理完成","全部任务均已完成",QSystemTrayIcon.MessageIcon.Information)
        else:
            self.tray.showMessage("fMRI处理完成","处理过程中出现错误",QSystemTrayIcon.MessageIcon.Critical)
    @pyqtSlot(int)
    def upgrade_progress(self,value:int):
        self.progress.setValue(value)
    @pyqtSlot()
    def busy_progress(self):
        self.progress.setRange(0,0)
        self.progress.setValue(0)
    def show_setting(self):
        setting=SettingDialog(self)
        setting.update_config_signal.connect(self.update_config)
        setting.resize(int(self.width()/1024*400),int(self.height()/768*150))
        setting.move(int(0.5*(self.width()-setting.width()))+self.pos().x(),int(0.5*(self.height()-setting.height()))+self.pos().y())
        setting.setStyleSheet("QDialog{background:#F3EAC2;border:none;border-radius:5px;}")
        setting.exec()
    def show_process_info_editor(self):
        editor=ProcessInfoEditor(self)
        editor.update_config_signal.connect(self.update_config)
        editor.resize(int(self.width()/1024*600),int(self.height()/768*300))
        editor.move(int(0.5*(self.width()-editor.width()))+self.pos().x(),int(0.5*(self.height()-editor.height()))+self.pos().y())
        editor.setStyleSheet("QDialog{background:#F3EAC2;border:none;border-radius:5px;}")
        editor.exec()
    @pyqtSlot()
    def close(self) -> bool:
        with open("config.json","r",encoding="utf-8") as p:
            config=json.load(p)
            config.update({"current_pos":[self.pos().x(),self.pos().y()]})
        with open("config.json","w",encoding="utf-8") as p:
            p.write(json.dumps(config,ensure_ascii=False,indent=4,sort_keys=True))
        self.tray.setVisible(False)
        self.tray.hide()
        return super().close()
    @pyqtSlot()
    def showMinimized(self) -> None:
        if self.conf["minimal_to_tray"]==True:
            self.setVisible(False)
            self.tray.setVisible(True)
            if self.need_tip==True:
                self.tray.showMessage(APPID,"已最小化到通知区域，双击可切换显示",QSystemTrayIcon.MessageIcon.Information)
                self.need_tip=False
        else:
            super().showMinimized()
    def mousePressEvent(self, event:QMouseEvent):
        self.logger.debug("触发鼠标按压事件")
        super().mousePressEvent(event)
        self.setFocus()
        self.m_flag=True
        if event.button()==Qt.MouseButton.LeftButton and self.isMaximized()==False and self.hasFocus()==True:
            self.old_pos=event.globalPosition() #获取鼠标相对窗口的位置
            self.logger.debug("已获取鼠标位置")
    def mouseMoveEvent(self, event:QMouseEvent):
        self.logger.debug("触发鼠标移动事件")
        super().mouseMoveEvent(event)
        if self.m_flag==True:
            delta_x=int(event.globalPosition().x()-self.old_pos.x())
            delta_y=int(event.globalPosition().y()-self.old_pos.y())
            self.move(self.x()+delta_x,self.y()+delta_y)#更改窗口位置
            self.logger.debug("已更改窗口位置")
            self.old_pos=event.globalPosition()
            self.setCursor(QCursor(Qt.CursorShape.SizeAllCursor))  #更改鼠标图标
    def mouseReleaseEvent(self, event:QMouseEvent):
        self.logger.debug("触发鼠标释放事件")
        super().mouseReleaseEvent(event)
        self.m_flag=False
        self.setCursor(QCursor(Qt.CursorShape.ArrowCursor))
if __name__=="__main__":
    app=QApplication(sys.argv)
    ui=UI()
    ui.show()
    exit(app.exec())
